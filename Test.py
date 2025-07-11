import openpyxl
from openpyxl.styles import Alignment
from datetime import datetime

class ProductionMapping:
    def __init__(self, source_col, target_sheet_name, name):
        self.source_col = source_col
        self.target_sheet_name = target_sheet_name
        self.name = name

class ArretMapping:
    def __init__(self, categorie, col_cible):
        self.categorie = categorie
        self.col_cible = col_cible

class S3Mapping:
    def __init__(self, chemin_s3, feuille_s3='AEE', col_date=4, col_valeur=69, ligne_debut=6):
        self.chemin_s3 = chemin_s3
        self.feuille_s3 = feuille_s3
        self.col_date = col_date
        self.col_valeur = col_valeur
        self.ligne_debut = ligne_debut
        self.date_to_value = {}
        self._charger_donnees()

    def _charger_donnees(self):
        wb_s3 = openpyxl.load_workbook(self.chemin_s3, data_only=True)
        ws_s3 = wb_s3[self.feuille_s3]
        for row in ws_s3.iter_rows(min_row=self.ligne_debut, max_row=ws_s3.max_row):
            date_cell = row[self.col_date - 1]
            value_cell = row[self.col_valeur - 1]
            if date_cell.value is not None and value_cell.value is not None:
                if isinstance(date_cell.value, datetime):
                    date_str = date_cell.value.strftime("%d/%m/%Y")
                else:
                    try:
                        date_obj = datetime.strptime(str(date_cell.value).strip(), "%d/%m/%Y")
                        date_str = date_obj.strftime("%d/%m/%Y")
                    except Exception:
                        date_str = str(date_cell.value).strip()
                self.date_to_value[date_str] = value_cell.value

    def get_value_for_date(self, date):
        if isinstance(date, datetime):
            date_str = date.strftime("%d/%m/%Y")
        else:
            try:
                date_obj = datetime.strptime(str(date).strip(), "%d/%m/%Y")
                date_str = date_obj.strftime("%d/%m/%Y")
            except Exception:
                date_str = str(date).strip()
        return self.date_to_value.get(date_str, None)

class S3AseMapping:
    def __init__(self, chemin_s3, feuille_s3='ASE', col_date=4, col_map=None, ligne_debut=6):
        if col_map is None:
            col_map = {'AL': 105, 'NL': 109, 'Ext': 113}
        self.chemin_s3 = chemin_s3
        self.feuille_s3 = feuille_s3
        self.col_date = col_date
        self.col_map = col_map
        self.ligne_debut = ligne_debut
        self.data = {k: {} for k in col_map}
        self._charger_donnees()

    def _charger_donnees(self):
        wb_s3 = openpyxl.load_workbook(self.chemin_s3, data_only=True)
        ws_s3 = wb_s3[self.feuille_s3]
        for row in ws_s3.iter_rows(min_row=self.ligne_debut, max_row=ws_s3.max_row):
            date_cell = row[self.col_date - 1]
            if date_cell.value is not None:
                if isinstance(date_cell.value, datetime):
                    date_str = date_cell.value.strftime("%d/%m/%Y")
                else:
                    try:
                        date_obj = datetime.strptime(str(date_cell.value).strip(), "%d/%m/%Y")
                        date_str = date_obj.strftime("%d/%m/%Y")
                    except Exception:
                        date_str = str(date_cell.value).strip()
                for tech, col in self.col_map.items():
                    value_cell = row[col - 1]
                    if value_cell.value is not None:
                        self.data[tech][date_str] = value_cell.value

    def get_value_for_date(self, tech, date):
        if isinstance(date, datetime):
            date_str = date.strftime("%d/%m/%Y")
        else:
            try:
                date_obj = datetime.strptime(str(date).strip(), "%d/%m/%Y")
                date_str = date_obj.strftime("%d/%m/%Y")
            except Exception:
                date_str = str(date).strip()
        return self.data[tech].get(date_str, None)

class JLNProcessor:
    def __init__(self, chemin_source1, chemin_source2, chemin_source3, chemin_cible):
        self.wb_source1 = openpyxl.load_workbook(chemin_source1, data_only=True)
        self.wb_source2 = openpyxl.load_workbook(chemin_source2, data_only=True)
        self.s3_mapping = S3Mapping(chemin_source3)
        self.s3_ase_mapping = S3AseMapping(chemin_source3)
        self.wb_cible = openpyxl.load_workbook(chemin_cible)
        self.ws_source_cap = self.wb_source1['CAP']
        self.ws_cible_ext = self.wb_cible['Ext']
        self.ws_cible_nl = self.wb_cible['NL']
        self.ws_cible_al = self.wb_cible['AL']
        self.ws_source2 = self.wb_source2['BD_Arrêts']
        self.production_mappings = [
            ProductionMapping(28, 'AL', 'AL'),
            ProductionMapping(29, 'NL', 'NL'),
            ProductionMapping(30, 'Ext', 'Ext')
        ]
        self.arret_mappings = [
            ArretMapping('planifies_maintenance', 7),
            ArretMapping('planifies_process', 8),
            ArretMapping('grande_revision', 9),
            ArretMapping('non_planifies_maintenance', 10),
            ArretMapping('externes', 12)
        ]

    def percent_str(self, val):
        try:
            return f"{round(float(val), 2)}%" if val is not None else None
        except Exception:
            return f"{val}%" if val is not None else None

    def trouver_derniere_ligne_date(self, feuille_cible):
        derniere_ligne = 1
        for row in range(2, feuille_cible.max_row + 1):
            valeur = feuille_cible.cell(row=row, column=1).value
            if valeur is not None and str(valeur).strip() != "":
                try:
                    if isinstance(valeur, datetime):
                        derniere_ligne = row
                    elif isinstance(valeur, str) and len(valeur.split('/')) == 3:
                        datetime.strptime(valeur, "%d/%m/%Y")
                        derniere_ligne = row
                except ValueError:
                    pass
        return derniere_ligne + 1

    def trouver_derniere_ligne_valide(self, feuille_source):
        lignes_vides_consecutives = 0
        derniere_ligne_valide = 3
        for row in range(4, feuille_source.max_row + 1):
            prod_al = feuille_source.cell(row=row, column=28).value
            prod_nl = feuille_source.cell(row=row, column=29).value
            prod_ext = feuille_source.cell(row=row, column=30).value
            if (prod_al is None or str(prod_al).strip() == "") and \
               (prod_nl is None or str(prod_nl).strip() == "") and \
               (prod_ext is None or str(prod_ext).strip() == ""):
                lignes_vides_consecutives += 1
            else:
                lignes_vides_consecutives = 0
                derniere_ligne_valide = row
            if lignes_vides_consecutives >= 3:
                break
        return derniere_ligne_valide

    def determiner_technologie(self, echelon):
        if echelon in ['I', 'J', 'V', 'W']:
            return 'NL'
        elif echelon in ['E', 'F', 'G', 'H']:
            return 'Ext'
        else:
            return 'AL'

    def mapper_imputation(self, imputation):
        mapping = {
            'PDI': 'externes',
            'PDM': 'non_planifies_maintenance',
            'PPLM': 'planifies_maintenance',
            'PPLP': 'planifies_process',
            'PPLM AF': 'grande_revision'
        }
        return mapping.get(imputation, 'autre')

    def traiter(self):
        # 1. Trouver la première ligne vide pour chaque feuille cible
        ligne_debut_ext = self.trouver_derniere_ligne_date(self.ws_cible_ext)
        ligne_debut_nl = self.trouver_derniere_ligne_date(self.ws_cible_nl)
        ligne_debut_al = self.trouver_derniere_ligne_date(self.ws_cible_al)
        # 2. Trouver la dernière ligne valide avec la condition d'arrêt
        derniere_ligne_valide = self.trouver_derniere_ligne_valide(self.ws_source_cap)
        # 3. Collecter toutes les dates existantes dans les trois feuilles cibles
        dates_existantes = set()
        for ws in [self.ws_cible_ext, self.ws_cible_nl, self.ws_cible_al]:
            for row in range(2, ws.max_row + 1):
                valeur = ws.cell(row=row, column=1).value
                if valeur is not None and str(valeur).strip() != "":
                    try:
                        if isinstance(valeur, datetime):
                            dates_existantes.add(valeur.strftime("%d/%m/%Y"))
                        elif isinstance(valeur, str) and len(valeur.split('/')) == 3:
                            datetime.strptime(valeur, "%d/%m/%Y")
                            dates_existantes.add(valeur)
                    except ValueError:
                        pass
        # 4. Collecter les dates du fichier source
        dates_a_ajouter = []
        for row in range(4, derniere_ligne_valide + 1):
            date_val = self.ws_source_cap.cell(row=row, column=1).value
            if date_val is None or str(date_val).strip() == "":
                continue
            try:
                if isinstance(date_val, datetime):
                    date_str = date_val.strftime("%d/%m/%Y")
                else:
                    date_val = datetime.strptime(str(date_val).strip(), "%d/%m/%Y")
                    date_str = date_val.strftime("%d/%m/%Y")
                if date_str not in dates_existantes:
                    dates_a_ajouter.append((row, date_str))
            except ValueError:
                continue
        # 5. Ajouter les nouvelles dates dans chaque feuille cible
        for mapping, ligne_debut in zip(self.production_mappings, [ligne_debut_al, ligne_debut_nl, ligne_debut_ext]):
            ws = self.wb_cible[mapping.target_sheet_name]
            current_row = ligne_debut
            for source_row, date_str in dates_a_ajouter:
                cell = ws.cell(row=current_row, column=1)
                cell.value = date_str
                cell.alignment = Alignment(horizontal='center', vertical='center')
                current_row += 1
        # 6. Copier les productions
        for mapping, ligne_debut in zip(self.production_mappings, [ligne_debut_al, ligne_debut_nl, ligne_debut_ext]):
            ws = self.wb_cible[mapping.target_sheet_name]
            current_row = ligne_debut
            for source_row, _ in dates_a_ajouter:
                prod_value = self.ws_source_cap.cell(row=source_row, column=mapping.source_col).value
                cell = ws.cell(row=current_row, column=6)
                cell.value = prod_value
                cell.alignment = Alignment(horizontal='center', vertical='center')
                current_row += 1
        # 7. Traiter les arrêts
        donnees_arrets = {}
        dates_valides = set(date_str for _, date_str in dates_a_ajouter)
        for row in range(2, self.ws_source2.max_row + 1):
            date_val = self.ws_source2.cell(row=row, column=1).value
            echelon = self.ws_source2.cell(row=row, column=4).value
            imputation = self.ws_source2.cell(row=row, column=7).value
            duree = self.ws_source2.cell(row=row, column=11).value
            if date_val is None or echelon is None or imputation is None or duree is None:
                continue
            try:
                if isinstance(date_val, datetime):
                    date_str = date_val.strftime("%d/%m/%Y")
                else:
                    date_val = datetime.strptime(str(date_val).strip(), "%d/%m/%Y")
                    date_str = date_val.strftime("%d/%m/%Y")
            except ValueError:
                continue
            if date_str not in dates_valides:
                continue
            technologie = self.determiner_technologie(str(echelon).strip())
            categorie = self.mapper_imputation(str(imputation).strip())
            if categorie == 'autre':
                continue
            try:
                duree_val = float(duree)
            except (ValueError, TypeError):
                continue
            if date_str not in donnees_arrets:
                donnees_arrets[date_str] = {}
            if technologie not in donnees_arrets[date_str]:
                donnees_arrets[date_str][technologie] = {m.categorie: 0 for m in self.arret_mappings}
            donnees_arrets[date_str][technologie][categorie] += duree_val
        # 8. Copier les arrêts dans les feuilles cibles
        for mapping, ligne_debut in zip(self.production_mappings, [ligne_debut_al, ligne_debut_nl, ligne_debut_ext]):
            ws = self.wb_cible[mapping.target_sheet_name]
            for idx, (_, date_str) in enumerate(dates_a_ajouter):
                row_cible = ligne_debut + idx
                if date_str in donnees_arrets and mapping.name in donnees_arrets[date_str]:
                    donnees_tech = donnees_arrets[date_str][mapping.name]
                    for arret_map in self.arret_mappings:
                        cell = ws.cell(row=row_cible, column=arret_map.col_cible)
                        cell.value = donnees_tech[arret_map.categorie]
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    for arret_map in self.arret_mappings:
                        cell = ws.cell(row=row_cible, column=arret_map.col_cible)
                        cell.value = 0
                        cell.alignment = Alignment(horizontal='center', vertical='center')
        # 9. Copier les valeurs du S3 dans la colonne 13 des feuilles cibles (matching sur la date)
        for feuille in ['AL', 'NL', 'Ext']:
            ws = self.wb_cible[feuille]
            for row in range(2, ws.max_row + 1):
                date_cell = ws.cell(row=row, column=1).value
                if date_cell is not None:
                    valeur_s3 = self.s3_mapping.get_value_for_date(date_cell)
                    if valeur_s3 is not None:
                        percent_val = self.percent_str(valeur_s3)
                        cell = ws.cell(row=row, column=13)
                        cell.value = percent_val
                        cell.alignment = Alignment(horizontal='center', vertical='center')
        # 10. Copier les valeurs de la feuille ASE du S3 dans la colonne 14 des feuilles cibles
        for feuille, col_s3 in zip(['AL', 'NL', 'Ext'], [105, 109, 113]):
            ws = self.wb_cible[feuille]
            for row in range(2, ws.max_row + 1):
                date_cell = ws.cell(row=row, column=1).value
                if date_cell is not None:
                    valeur_ase = self.s3_ase_mapping.get_value_for_date(feuille, date_cell)
                    if valeur_ase is not None:
                        percent_val = self.percent_str(valeur_ase)
                        cell = ws.cell(row=row, column=14)
                        cell.value = percent_val
                        cell.alignment = Alignment(horizontal='center', vertical='center')
        # 11. Centrer toutes les autres données recopiées (dates, production, arrêts)
        for feuille in ['AL', 'NL', 'Ext']:
            ws = self.wb_cible[feuille]
            for row in range(2, ws.max_row + 1):
                for col in [1, 6, 7, 8, 9, 10, 11, 12]:
                    cell = ws.cell(row=row, column=col)
                    if cell.value is not None:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
        # 12. Sauvegarder
        fichier_sortie = 'Fichier Prétraité JLN.xlsx'
        self.wb_cible.save(fichier_sortie)
        print("\n✅ ----------------------------------------------------SUCCES--------------------------------------------")
        return fichier_sortie

def traiter_fichiers_JLN(chemin_source1, chemin_source2, chemin_source3, chemin_cible):
    processor = JLNProcessor(chemin_source1, chemin_source2, chemin_source3, chemin_cible)
    return processor.traiter()

if __name__ == "__main__":
    fichier_s1 = 'Fichier Source S1 Prod JLN.xlsx'
    fichier_s2 = 'Fichier Source S2 Arrêts JLN.xlsx'
    fichier_s3 = 'Fichier Source S3 Titres JLN.xlsx'
    fichier_cible = 'Fichier Cible JLN.xlsx'
    fichier_resultat = traiter_fichiers_JLN(fichier_s1, fichier_s2, fichier_s3, fichier_cible)
    print(f"Fichier de sortie généré : {fichier_resultat}")
