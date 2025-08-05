import streamlit as st
import tempfile
import os
import openpyxl
from main2 import traiter_fichiers_MP1
from Test import traiter_fichiers_JLN
from mainJFC2 import traiter_jfc2
from mainJFC4 import traiter_jfc4
from mainJFC5 import traiter_fichiers_JFC5
from mainJFC1 import traiter_fichiers_pretraitement

# --- Personnalisation du thème via CSS ---
st.markdown(
    """
    <style>
    .stButton>button {
        background-color: #0A6EBD;
        color: white;
        font-weight: bold;
        border-radius: 8px;
        padding: 0.5em 2em;
        margin-top: 1em;
    }
    .stDownloadButton>button {
        background-color: #198754;
        color: white;
        font-weight: bold;
        border-radius: 8px;
        padding: 0.5em 2em;
    }
    .stRadio>div>label {
        font-size: 1.1em;
        font-weight: 500;
    }
    </style>
    """,
    unsafe_allow_html=True
)

# --- Logo (optionnel) ---
# Place un fichier 'logo.png' dans le dossier du projet si tu veux l'afficher
if os.path.exists("logo.png"):
    st.image("logo.png", width=120)

# --- Titres et sous-titres ---
st.title("Prétraitement de données brutes")
st.markdown("<h4 style='color:#0A6EBD;'>Automatisez le prétraitement de vos fichiers Excel MP1, JLN, JFC1, JFC2, JFC4, JFC5</h4>", unsafe_allow_html=True)
st.markdown("---")

st.write("""
Bienvenue sur l'interface de prétraitement de données brutes. Sélectionnez le module à traiter, chargez vos fichiers, puis lancez le traitement. Le fichier prétraité sera généré automatiquement.
""")

# --- Choix du module ---
modules = {
    "MP1": "Prétraitement MP1",
    "JLN": "Prétraitement JLN",
    "JFC1": "Prétraitement JFC1",
    "JFC2": "Prétraitement JFC2",
    "JFC4": "Prétraitement JFC4",
    "JFC5": "Prétraitement JFC5"
}
module = st.selectbox("Choisissez le module à traiter :", list(modules.keys()), format_func=lambda x: modules[x])

if module == "MP1":
    st.markdown("## 🔵 Prétraitement MP1")
    st.info("Veuillez charger le fichier source et le fichier cible MP1.")
    fichier_source = st.file_uploader("Fichier source MP1", type=["xlsx"], key="src_mp1")
    fichier_cible = st.file_uploader("Fichier cible MP1", type=["xlsx"], key="cible_mp1")
    if fichier_source and fichier_cible:
        col1, col2, col3 = st.columns([1,2,1])
        with col2:
            if st.button("Lancer le traitement MP1"):
                with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_source, \
                     tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_cible:
                    tmp_source.write(fichier_source.read())
                    tmp_cible.write(fichier_cible.read())
                    tmp_source.flush()
                    tmp_cible.flush()
                    with st.spinner("Traitement en cours... Merci de patienter."):
                        try:
                            fichier_sortie = traiter_fichiers_MP1(tmp_source.name, tmp_cible.name)
                            if fichier_sortie:
                                with open(fichier_sortie, "rb") as f:
                                    st.success("Traitement terminé ! Téléchargez le fichier prétraité ci-dessous.")
                                    st.download_button(
                                        label="Télécharger le fichier prétraité",
                                        data=f,
                                        file_name="Fichier Prétraité MP1.xlsx",
                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                    )
                            else:
                                st.warning("Le traitement n'a pas pu être effectué. Vérifiez vos fichiers.")
                        except KeyError as e:
                            if "xl/drawings/NULL" in str(e):
                                st.error("Erreur : Le fichier Excel contient une référence à un dessin/image corrompu ou manquant. Ouvrez le fichier dans Excel, supprimez tous les objets graphiques (images, graphiques, etc.), puis réenregistrez-le.")
                            else:
                                st.error(f"Erreur inattendue : {e}")
                        except Exception as e:
                            st.error(f"Erreur inattendue : {e}")

elif module == "JLN":
    st.markdown("## 🟢 Prétraitement JLN")
    st.info("Veuillez charger les trois fichiers source (Production, Arrêts, Titres) et le fichier cible JLN.")
    fichier_source1 = st.file_uploader("Fichier source 1 (Production)", type=["xlsx"], key="src1_jln")
    fichier_source2 = st.file_uploader("Fichier source 2 (Arrêts)", type=["xlsx"], key="src2_jln")
    fichier_source3 = st.file_uploader("Fichier source 3 (Titres)", type=["xlsx"], key="src3_jln")
    fichier_cible = st.file_uploader("Fichier cible JLN", type=["xlsx"], key="cible_jln")
    if fichier_source1 and fichier_source2 and fichier_source3 and fichier_cible:
        col1, col2, col3 = st.columns([1,2,1])
        with col2:
            if st.button("Lancer le traitement JLN"):
                with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_s1, \
                     tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_s2, \
                     tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_s3, \
                     tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_cible:
                    tmp_s1.write(fichier_source1.read())
                    tmp_s2.write(fichier_source2.read())
                    tmp_s3.write(fichier_source3.read())
                    tmp_cible.write(fichier_cible.read())
                    tmp_s1.flush()
                    tmp_s2.flush()
                    tmp_s3.flush()
                    tmp_cible.flush()
                    with st.spinner("Traitement en cours... Merci de patienter."):
                        try:
                            fichier_sortie = traiter_fichiers_JLN(tmp_s1.name, tmp_s2.name, tmp_s3.name, tmp_cible.name)
                            if fichier_sortie:
                                with open(fichier_sortie, "rb") as f:
                                    st.success("Traitement terminé ! Téléchargez le fichier prétraité ci-dessous.")
                                    st.download_button(
                                        label="Télécharger le fichier prétraité",
                                        data=f,
                                        file_name="Fichier Prétraité JLN.xlsx",
                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                    )
                            else:
                                st.warning("Le traitement n'a pas pu être effectué. Vérifiez vos fichiers.")
                        except Exception as e:
                            st.error(f"Erreur inattendue : {e}")

elif module == "JFC1":
    st.markdown("## 🟡 Prétraitement JFC1")
    st.info("Veuillez charger le fichier source et le fichier cible JFC1.")
    fichier_source = st.file_uploader("Fichier source JFC1", type=["xlsx"], key="src_jfc1")
    fichier_cible = st.file_uploader("Fichier cible JFC1", type=["xlsx"], key="cible_jfc1")
    if fichier_source and fichier_cible:
        col1, col2, col3 = st.columns([1,2,1])
        with col2:
            if st.button("Lancer le traitement JFC1"):
                with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_source, \
                     tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_cible:
                    tmp_source.write(fichier_source.read())
                    tmp_cible.write(fichier_cible.read())
                    tmp_source.flush()
                    tmp_cible.flush()
                    with st.spinner("Traitement en cours... Merci de patienter."):
                        try:
                            fichier_sortie = traiter_fichiers_pretraitement(tmp_source.name, tmp_cible.name)
                            if fichier_sortie:
                                with open(fichier_sortie, "rb") as f:
                                    st.success("Traitement terminé ! Téléchargez le fichier prétraité ci-dessous.")
                                    st.download_button(
                                        label="Télécharger le fichier prétraité",
                                        data=f,
                                        file_name="Fichier Prétraité JFC1.xlsx",
                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                    )
                            else:
                                st.warning("Le traitement n'a pas pu être effectué. Vérifiez vos fichiers.")
                        except Exception as e:
                            st.error(f"Erreur inattendue : {e}")

elif module == "JFC2":
    st.markdown("## 🟣 Prétraitement JFC2")
    st.info("Veuillez charger le fichier source et le fichier cible JFC2.")
    fichier_source = st.file_uploader("Fichier source JFC2", type=["xlsx"], key="src_jfc2")
    fichier_cible = st.file_uploader("Fichier cible JFC2", type=["xlsx"], key="cible_jfc2")
    if fichier_source and fichier_cible:
        col1, col2, col3 = st.columns([1,2,1])
        with col2:
            if st.button("Lancer le traitement JFC2"):
                with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_source, \
                     tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_cible, \
                     tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_sortie:
                    tmp_source.write(fichier_source.read())
                    tmp_cible.write(fichier_cible.read())
                    tmp_source.flush()
                    tmp_cible.flush()
                    with st.spinner("Traitement en cours... Merci de patienter."):
                        try:
                            traiter_jfc2(tmp_source.name, tmp_cible.name, tmp_sortie.name)
                            with open(tmp_sortie.name, "rb") as f:
                                st.success("Traitement terminé ! Téléchargez le fichier prétraité ci-dessous.")
                                st.download_button(
                                    label="Télécharger le fichier prétraité",
                                    data=f,
                                    file_name="Fichier Prétraité JFC2.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                )
                        except Exception as e:
                            st.error(f"Erreur inattendue : {e}")

elif module == "JFC4":
    st.markdown("## 🟤 Prétraitement JFC4")
    st.info("Veuillez charger le fichier source et le fichier cible JFC4.")
    fichier_source = st.file_uploader("Fichier source JFC4", type=["xlsx"], key="src_jfc4")
    fichier_cible = st.file_uploader("Fichier cible JFC4", type=["xlsx"], key="cible_jfc4")
    if fichier_source and fichier_cible:
        col1, col2, col3 = st.columns([1,2,1])
        with col2:
            if st.button("Lancer le traitement JFC4"):
                with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_source, \
                     tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_cible, \
                     tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_sortie:
                    tmp_source.write(fichier_source.read())
                    tmp_cible.write(fichier_cible.read())
                    tmp_source.flush()
                    tmp_cible.flush()
                    with st.spinner("Traitement en cours... Merci de patienter."):
                        try:
                            traiter_jfc4(tmp_source.name, tmp_cible.name, tmp_sortie.name)
                            with open(tmp_sortie.name, "rb") as f:
                                st.success("Traitement terminé ! Téléchargez le fichier prétraité ci-dessous.")
                                st.download_button(
                                    label="Télécharger le fichier prétraité",
                                    data=f,
                                    file_name="Fichier Prétraité JFC4.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                )
                        except Exception as e:
                            st.error(f"Erreur inattendue : {e}")

elif module == "JFC5":
    st.markdown("## 🟠 Prétraitement JFC5")
    st.info("Veuillez charger les deux fichiers source et le fichier cible JFC5.")
    fichier_source1 = st.file_uploader("Fichier source 1 JFC5", type=["xlsx"], key="src1_jfc5")
    fichier_source2 = st.file_uploader("Fichier source 2 JFC5", type=["xlsx"], key="src2_jfc5")
    fichier_cible = st.file_uploader("Fichier cible JFC5", type=["xlsx"], key="cible_jfc5")
    if fichier_source1 and fichier_source2 and fichier_cible:
        col1, col2, col3 = st.columns([1,2,1])
        with col2:
            if st.button("Lancer le traitement JFC5"):
                with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_s1, \
                     tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_s2, \
                     tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_cible:
                    tmp_s1.write(fichier_source1.read())
                    tmp_s2.write(fichier_source2.read())
                    tmp_cible.write(fichier_cible.read())
                    tmp_s1.flush()
                    tmp_s2.flush()
                    tmp_cible.flush()
                    with st.spinner("Traitement en cours... Merci de patienter."):
                        try:
                            fichier_sortie = traiter_fichiers_JFC5(tmp_s1.name, tmp_s2.name, tmp_cible.name)
                            if fichier_sortie:
                                with open(fichier_sortie, "rb") as f:
                                    st.success("Traitement terminé ! Téléchargez le fichier prétraité ci-dessous.")
                                    st.download_button(
                                        label="Télécharger le fichier prétraité",
                                        data=f,
                                        file_name="Fichier Prétraité JFC5.xlsx",
                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                    )
                            else:
                                st.warning("Le traitement n'a pas pu être effectué. Vérifiez vos fichiers.")
                        except Exception as e:
                            st.error(f"Erreur inattendue : {e}")

st.markdown("---")
st.header("📁 Générer le Fichier Global Jorf")
st.write("Chargez un ou plusieurs fichiers prétraités JLN,JFC1, JFC2, JFC4, JFC5, puis cliquez sur le bouton pour générer un fichier global avec un onglet par entité.")

col_jln, col_jfc1, col_jfc2, col_jfc4, col_jfc5 = st.columns(5)
with col_jln:
    fichier_jln = st.file_uploader("JLN", type=["xlsx"], key="global_jln")
with col_jfc1:
    fichier_jfc1 = st.file_uploader("JFC1", type=["xlsx"], key="global_jfc1")
with col_jfc2:
    fichier_jfc2 = st.file_uploader("JFC2", type=["xlsx"], key="global_jfc2")
with col_jfc4:
    fichier_jfc4 = st.file_uploader("JFC4", type=["xlsx"], key="global_jfc4")
with col_jfc5:
    fichier_jfc5 = st.file_uploader("JFC5", type=["xlsx"], key="global_jfc5")

fichiers = {
    "JLN": fichier_jln,
    "JFC1": fichier_jfc1,
    "JFC2": fichier_jfc2,
    "JFC4": fichier_jfc4,
    "JFC5": fichier_jfc5
}

if any(fichiers.values()):
    if st.button("Générer le fichier global Jorf"):
        with st.spinner("Génération du fichier global en cours..."):
            from io import BytesIO
            import openpyxl
            from copy import copy
            def copy_worksheet(source_ws, target_ws):
                # Copie des valeurs et styles
                for row in source_ws.iter_rows():
                    for cell in row:
                        new_cell = target_ws.cell(row=cell.row, column=cell.col_idx, value=cell.value)
                        if cell.has_style:
                            new_cell.font = copy(cell.font)
                            new_cell.border = copy(cell.border)
                            new_cell.fill = copy(cell.fill)
                            new_cell.number_format = copy(cell.number_format)
                            new_cell.protection = copy(cell.protection)
                            new_cell.alignment = copy(cell.alignment)
                # Largeurs de colonnes
                for col_letter, dim in source_ws.column_dimensions.items():
                    target_ws.column_dimensions[col_letter].width = dim.width
                # Hauteurs de lignes
                for row_idx, dim in source_ws.row_dimensions.items():
                    target_ws.row_dimensions[row_idx].height = dim.height
                # Merges
                for merged_range in source_ws.merged_cells.ranges:
                    target_ws.merge_cells(str(merged_range))
            wb_global = openpyxl.Workbook()
            # Supprimer la feuille par défaut
            default_sheet = wb_global.active
            wb_global.remove(default_sheet)
            for entite, fichier in fichiers.items():
                if fichier:
                    wb_entite = openpyxl.load_workbook(fichier, data_only=False)
                    if entite == "JLN":
                        for sheet_name in wb_entite.sheetnames:
                            ws_entite = wb_entite[sheet_name]
                            ws_global = wb_global.create_sheet(title=sheet_name)
                            copy_worksheet(ws_entite, ws_global)
                    else:
                        ws_entite = wb_entite.active
                        ws_global = wb_global.create_sheet(title=entite)
                        copy_worksheet(ws_entite, ws_global)
            output = BytesIO()
            wb_global.save(output)
            output.seek(0)
            st.success("Fichier global généré ! Téléchargez-le ci-dessous.")
            st.download_button(
                label="Télécharger le fichier global Jorf",
                data=output,
                file_name="Fichier Global Jorf.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )



