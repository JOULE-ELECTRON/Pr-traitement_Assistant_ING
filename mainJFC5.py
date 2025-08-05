import openpyxl
from openpyxl.styles import Alignment
from datetime import datetime

class Mapping:
    def __init__(self, source_col, target_col, name):
        self.source_col = source_col
        self.target_col = target_col
        self.name = name

class JFC5Processor:
    def __init__(self, chemin_source1, chemin_source2, chemin_cible):
        self.wb_source1 = openpyxl.load_workbook(chemin_source1, data_only=True)
        self.ws_source1 = self.wb_source1.active
        self.source1_max_row = self.ws_source1.max_row
        self.wb_source2 = openpyxl.load_workbook(chemin_source2, data_only=True)
        self.ws_source2 = self.wb_source2.active
        self.source2_max_row = self.ws_source2.max_row
        self.wb_cible = openpyxl.load_workbook(chemin_cible)
        try:
            self.ws_cible = self.wb_cible['JFC5']
        except KeyError:
            raise ValueError("La feuille 'JFC5' n'existe pas dans le fichier cible.")
        # Mapping S1 (1-based)
        self.date_col_source1 = 2
        self.date_col_cible = 1
        self.production_mapping = Mapping(5, 6, "Production")
        self.arret_mappings = [
            Mapping(7, 8, "Temps d'arrêts planifiés process (h)"),
            Mapping(8, 7, "Temps d'arrêts planifiés maintenance (h)"),
            Mapping(9, 9, "Temps grande révision (h)"),
            Mapping(10, 11, "Temps d'arrêts non planifiés process (h)"),
            Mapping(11, 10, "Temps d'arrêts non planifiés maintenance (h)"),
            Mapping(12, 12, "Temps d'arrêts externes (h)")
        ]
        # Mapping S2 titres
        self.titre_mappings = [
            Mapping(4, 13, "ACP 29% > Titre"),
            Mapping(7, 14, "ACP54% > Titre")
        ]

    def percent_str(self, val):
        try:
            return f"{round(float(val), 2)}%" if val is not None else None
        except Exception:
            return f"{val}%" if val is not None else None

    def trouver_derniere_ligne_date(self):
        derniere_ligne = 1
        for row in range(2, self.ws_cible.max_row + 1):
            valeur = self.ws_cible.cell(row=row, column=self.date_col_cible).value
            if valeur is not None and str(valeur).strip() != "":
                try:
                    if isinstance(valeur, datetime):
                        derniere_ligne = row
                    elif isinstance(valeur, str) and len(valeur.split('/')) == 3:
                        datetime.strptime(valeur, "%d/%m/%Y")
                        derniere_ligne = row
                except ValueError:
                    pass
        return derniere_ligne + 1

    def collecter_dates_existantes(self):
        dates_existantes = set()
        for row in range(2, self.ws_cible.max_row + 1):
            valeur = self.ws_cible.cell(row=row, column=self.date_col_cible).value
            if valeur is not None and str(valeur).strip() != "":
                try:
                    if isinstance(valeur, datetime):
                        dates_existantes.add(valeur.strftime("%d/%m/%Y"))
                    elif isinstance(valeur, str) and len(valeur.split('/')) == 3:
                        datetime.strptime(valeur, "%d/%m/%Y")
                        dates_existantes.add(valeur)
                except ValueError:
                    pass
        return dates_existantes

    def collecter_titres(self):
        # Récupère les titres de S2 sous forme {date_str: {"acp29": val, "acp54": val}}
        titres = {}
        for row in range(3, self.source2_max_row + 1):
            date_val = self.ws_source2.cell(row=row, column=2).value
            if date_val is None or str(date_val).strip() == "":
                continue
            try:
                if isinstance(date_val, datetime):
                    date_str = date_val.strftime("%d/%m/%Y")
                else:
                    date_val = datetime.strptime(str(date_val).strip(), "%d/%m/%Y")
                    date_str = date_val.strftime("%d/%m/%Y")
            except Exception:
                continue
            acp29 = self.ws_source2.cell(row=row, column=4).value
            acp54 = self.ws_source2.cell(row=row, column=7).value
            titres[date_str] = {"acp29": acp29, "acp54": acp54}
        return titres

    def traiter(self):
        ligne_debut = self.trouver_derniere_ligne_date()
        dates_existantes = self.collecter_dates_existantes()
        titres = self.collecter_titres()
        dates_a_ajouter = []  # (row_source, date_str)
        for row in range(3, self.source1_max_row + 1):
            date_val = self.ws_source1.cell(row=row, column=self.date_col_source1).value
            if date_val is None or str(date_val).strip() == "":
                continue
            try:
                if isinstance(date_val, datetime):
                    date_str = date_val.strftime("%d/%m/%Y")
                else:
                    date_val = datetime.strptime(str(date_val).strip(), "%d/%m/%Y")
                    date_str = date_val.strftime("%d/%m/%Y")
                if date_str not in dates_existantes:
                    dates_a_ajouter.append((row, date_str))
            except Exception:
                continue
        current_row = ligne_debut
        for row_source, date_str in dates_a_ajouter:
            # 1. Date
            cell = self.ws_cible.cell(row=current_row, column=self.date_col_cible)
            cell.value = date_str
            cell.alignment = Alignment(horizontal='center', vertical='center')
            # 2. Production
            prod_value = self.ws_source1.cell(row=row_source, column=self.production_mapping.source_col).value
            cell = self.ws_cible.cell(row=current_row, column=self.production_mapping.target_col)
            cell.value = prod_value
            cell.alignment = Alignment(horizontal='center', vertical='center')
            # 3. Arrêts
            for arret_map in self.arret_mappings:
                arret_value = self.ws_source1.cell(row=row_source, column=arret_map.source_col).value
                cell = self.ws_cible.cell(row=current_row, column=arret_map.target_col)
                cell.value = arret_value
                cell.alignment = Alignment(horizontal='center', vertical='center')
            # 4. Titres (si présents pour la date)
            if date_str in titres:
                titre_acp29 = titres[date_str]['acp29']
                titre_acp54 = titres[date_str]['acp54']
                percent_acp29 = self.percent_str(titre_acp29)
                percent_acp54 = self.percent_str(titre_acp54)
                values = [(13, percent_acp29), (14, percent_acp54)]
                for col, val in values:
                    cell = self.ws_cible.cell(row=current_row, column=col)
                    cell.value = val
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1
        fichier_sortie = 'Fichier Prétraité JFC5.xlsx'
        self.wb_cible.save(fichier_sortie)
        print("\n✅ ----------------------------------------------------SUCCES--------------------------------------------")
        return fichier_sortie

def traiter_fichiers_JFC5(chemin_source1, chemin_source2, chemin_cible):
    processor = JFC5Processor(chemin_source1, chemin_source2, chemin_cible)
    return processor.traiter()

if __name__ == "__main__":
    fichier_s1 = 'Fichier Source S1 JFC5.xlsx'
    fichier_s2 = 'Fichier Source S2 JFC5.xlsx'
    fichier_cible = 'Fichier Cible Jorf.xlsx'
    fichier_resultat = traiter_fichiers_JFC5(fichier_s1, fichier_s2, fichier_cible)
    print(f"Fichier de sortie généré : {fichier_resultat}")
