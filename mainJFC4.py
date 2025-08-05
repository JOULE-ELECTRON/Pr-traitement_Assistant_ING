import openpyxl
from datetime import datetime
from openpyxl.styles import Alignment

class JFC4SourceReader:
    def __init__(self, chemin_source):
        self.wb = openpyxl.load_workbook(chemin_source, data_only=True)
        self.ws = self.wb.active

    def get_dates(self):
        # Les dates sont sur la ligne 2, à partir de la colonne 3
        return [self.ws.cell(row=2, column=col).value for col in range(3, self.ws.max_column + 1)]

    def get_production(self):
        # Production > ACP 54NCl sur la ligne 5, à partir de la colonne 3
        return [self.ws.cell(row=5, column=col).value for col in range(3, self.ws.max_column + 1)]

    def get_arrets(self):
        # Arrêts sur les lignes 44 à 49, à partir de la colonne 3
        return {
            'planifies_process': [self.ws.cell(row=44, column=col).value for col in range(3, self.ws.max_column + 1)],
            'planifies_maintenance': [self.ws.cell(row=45, column=col).value for col in range(3, self.ws.max_column + 1)],
            'grande_revision': [self.ws.cell(row=46, column=col).value for col in range(3, self.ws.max_column + 1)],
            'non_planifies_process': [self.ws.cell(row=47, column=col).value for col in range(3, self.ws.max_column + 1)],
            'non_planifies_maintenance': [self.ws.cell(row=48, column=col).value for col in range(3, self.ws.max_column + 1)],
            'externes': [self.ws.cell(row=49, column=col).value for col in range(3, self.ws.max_column + 1)],
        }

    def get_titres(self):
        # Titres sur les lignes 30 (ACP 29) et 31 (ACP 54), à partir de la colonne 3
        return {
            'acp29': [self.ws.cell(row=30, column=col).value for col in range(3, self.ws.max_column + 1)],
            'acp54': [self.ws.cell(row=31, column=col).value for col in range(3, self.ws.max_column + 1)],
        }

class JFC4CibleWriter:
    def __init__(self, chemin_cible):
        self.wb_cible = openpyxl.load_workbook(chemin_cible)
        try:
            self.ws_cible = self.wb_cible['JFC4']
        except KeyError:
            raise ValueError("La feuille 'JFC4' n'existe pas dans le fichier cible.")

    def get_existing_dates(self):
        # Les dates sont en colonne 1, à partir de la ligne 2
        dates = set()
        for row in range(2, self.ws_cible.max_row + 1):
            val = self.ws_cible.cell(row=row, column=1).value
            if val:
                if isinstance(val, datetime):
                    dates.add(val.strftime("%d/%m/%Y"))
                else:
                    dates.add(str(val))
        return dates

    def find_first_empty_row(self):
        for row in range(2, self.ws_cible.max_row + 2):
            if not self.ws_cible.cell(row=row, column=1).value:
                return row
        return self.ws_cible.max_row + 1

    def write_row(self, row, date, production, arrets, titres):
        # Colonnes à remplir : 1 (date), 6 (prod), 7-12 (arrêts), 13-14 (titres)
        values = [
            (1, date),
            (6, production),
            (7, arrets['planifies_maintenance']),
            (8, arrets['planifies_process']),
            (9, arrets['grande_revision']),
            (10, arrets['non_planifies_maintenance']),
            (11, arrets['non_planifies_process']),
            (12, arrets['externes'])
        ]
        # Conversion en pourcentage sous forme de chaîne avec 2 décimales
        titre_acp29 = titres['acp29']
        titre_acp54 = titres['acp54']
        def percent_str(val):
            try:
                return f"{round(float(val), 2)}%" if val is not None else None
            except Exception:
                return f"{val}%" if val is not None else None
        percent_acp29 = percent_str(titre_acp29)
        percent_acp54 = percent_str(titre_acp54)
        values.append((13, percent_acp29))
        values.append((14, percent_acp54))
        for col, val in values:
            cell = self.ws_cible.cell(row=row, column=col)
            cell.value = val
            cell.alignment = Alignment(horizontal='center', vertical='center')

    def save(self, chemin_sortie):
        self.wb_cible.save(chemin_sortie)

def traiter_jfc4(chemin_source, chemin_cible, chemin_sortie):
    reader = JFC4SourceReader(chemin_source)
    writer = JFC4CibleWriter(chemin_cible)

    dates = reader.get_dates()
    productions = reader.get_production()
    arrets = reader.get_arrets()
    titres = reader.get_titres()

    existing_dates = writer.get_existing_dates()

    for idx, date in enumerate(dates):
        if date is None:
            continue
        if isinstance(date, datetime):
            date_str = date.strftime("%d/%m/%Y")
        else:
            date_str = str(date)
        if date_str in existing_dates or not date_str.strip() or date_str.lower() == 'none':
            continue
        row = writer.find_first_empty_row()
        writer.write_row(
            row,
            date_str,
            productions[idx],
            {k: v[idx] for k, v in arrets.items()},
            {k: v[idx] for k, v in titres.items()}
        )
        existing_dates.add(date_str)

    writer.save(chemin_sortie)
    print(f"✅ Fichier prétraité généré : {chemin_sortie}")

# Exemple d'appel (à adapter selon tes fichiers) :
# traiter_jfc2('Fichier Source JFC2.xlsx', 'Fichier Cible Jorf.xlsx', 'Fichier Prétraité JFC2.xlsx')

if __name__ == "__main__":
    fichier_source = "Fichier Source JFC4.xlsx"
    fichier_cible = "Fichier Cible Jorf.xlsx"
    fichier_sortie = "Fichier Prétraité JFC4.xlsx"
    traiter_jfc4(fichier_source, fichier_cible, fichier_sortie)
