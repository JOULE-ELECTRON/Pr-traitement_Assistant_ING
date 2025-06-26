import openpyxl
from datetime import datetime, timedelta
import re

class ColonneMapping:
    def __init__(self, feuille_source, col_cible, col_source=None, colonnes_source=None, ligne_debut_source=None, ligne_debut_cible=None, type=None):
        self.feuille_source = feuille_source
        self.col_cible = col_cible
        self.col_source = col_source
        self.colonnes_source = colonnes_source
        self.ligne_debut_source = ligne_debut_source
        self.ligne_debut_cible = ligne_debut_cible
        self.type = type

def traiter_fichiers_MP1(chemin_source, chemin_cible):
    """
    Traite les fichiers Excel pour le calcul des temps d'arrêt et de production.
    
    Args:
        chemin_source (str): Chemin vers le fichier source MP1
        chemin_cible (str): Chemin vers le fichier cible MP1
        
    Returns:
        str: Chemin vers le fichier de sortie généré
    """
    # Charger les fichiers Excel
    wb_source = openpyxl.load_workbook(chemin_source, data_only=False)
    wb_cible = openpyxl.load_workbook(chemin_cible)

    ws_cible = wb_cible.active

    # Fonction pour trouver l'indice de la colonne cible dans le fichier cible
    def trouver_colonne(feuille, nom_colonne):
        for col in range(1, feuille.max_column + 1):
            valeur = feuille.cell(row=1, column=col).value
            if valeur and str(valeur).strip().lower() == nom_colonne.strip().lower():
                return col
        raise ValueError(f"Colonne '{nom_colonne}' non trouvée.")

    # Fonction pour calculer la date suivante
    def calculer_date_suivante(date_str, jours=1):
        try:
            # Si c'est un objet datetime, on le convertit en string
            if isinstance(date_str, datetime):
                date_str = date_str.strftime("%d/%m/%Y")
            # On parse la date
            date = datetime.strptime(date_str, "%d/%m/%Y")
            # On ajoute le nombre de jours spécifié
            date_suivante = date + timedelta(days=jours)
            # On retourne au format JJ/MM/AAAA
            return date_suivante.strftime("%d/%m/%Y")
        except:
            return date_str

    # Fonction pour trouver la dernière ligne avec une date dans la colonne 1
    def trouver_derniere_ligne_date(feuille_cible):
        derniere_ligne = 1
        for row in range(1, feuille_cible.max_row + 1):
            valeur = feuille_cible.cell(row=row, column=1).value
            if valeur is not None and str(valeur).strip() != "":
                # Vérifier si c'est une date (format JJ/MM/AAAA)
                try:
                    if isinstance(valeur, datetime):
                        derniere_ligne = row
                    elif isinstance(valeur, str) and len(valeur.split('/')) == 3:
                        datetime.strptime(valeur, "%d/%m/%Y")
                        derniere_ligne = row
                except ValueError:
                    pass  # Ce n'est pas une date, on continue
        return derniere_ligne

    # Fonction pour détecter automatiquement la dernière date valide dans le fichier source
    def detecter_derniere_date_source(feuille_source, col_source, ligne_debut_source):
        derniere_date = None
        derniere_ligne_date = None
        
        # Parcourir la colonne source pour trouver le dernier "Total" suivi de deux lignes vides
        dernier_total_ligne = None
        for row in range(ligne_debut_source, feuille_source.max_row + 1):
            valeur = feuille_source.cell(row=row, column=col_source).value
            
            if valeur == "Total":
                # Vérifier si les deux lignes suivantes sont vides
                ligne_suiv1 = feuille_source.cell(row=row + 1, column=col_source).value
                ligne_suiv2 = feuille_source.cell(row=row + 2, column=col_source).value
                
                if (ligne_suiv1 is None or str(ligne_suiv1).strip() == "") and (ligne_suiv2 is None or str(ligne_suiv2).strip() == ""):
                    dernier_total_ligne = row
                    print(f"Dernier Total trouvé à la ligne {row}")
        
        # Maintenant, trouver la dernière date valide avant ce dernier Total
        if dernier_total_ligne:
            for row in range(ligne_debut_source, dernier_total_ligne):
                valeur = feuille_source.cell(row=row, column=col_source).value
                if valeur is not None and str(valeur).strip() != "":
                    # Vérifier si c'est une date
                    try:
                        if isinstance(valeur, datetime):
                            derniere_date = valeur.strftime("%d/%m/%Y")
                            derniere_ligne_date = row
                        elif isinstance(valeur, str) and len(valeur.split('/')) == 3:
                            datetime.strptime(valeur, "%d/%m/%Y")
                            derniere_date = valeur
                            derniere_ligne_date = row
                    except ValueError:
                        pass  # Ce n'est pas une date, on continue
        
        return derniere_date, derniere_ligne_date, dernier_total_ligne

    # Trouver automatiquement la ligne de départ dans le fichier cible
    ligne_debut_cible_auto = trouver_derniere_ligne_date(ws_cible) + 1
    print(f"Ligne de départ automatiquement détectée : {ligne_debut_cible_auto}")

    # Liste des colonnes à copier
    colonnes_a_copier = [
        ColonneMapping("Synthèse", 1, col_source=2, ligne_debut_source=10, ligne_debut_cible=ligne_debut_cible_auto),
        ColonneMapping("Synthèse", 6, col_source=4, ligne_debut_source=10, ligne_debut_cible=ligne_debut_cible_auto),
        ColonneMapping("Synthèse", 19, col_source=7, ligne_debut_source=10, ligne_debut_cible=ligne_debut_cible_auto),
        ColonneMapping("Synthèse", 32, col_source=10, ligne_debut_source=10, ligne_debut_cible=ligne_debut_cible_auto),
        ColonneMapping("Synthèse", 45, col_source=13, ligne_debut_source=10, ligne_debut_cible=ligne_debut_cible_auto),
        ColonneMapping("Suivi des arrêts", 7, col_source=3, ligne_debut_source=11, ligne_debut_cible=ligne_debut_cible_auto),
        ColonneMapping("Suivi des arrêts", 20, col_source=29, ligne_debut_source=11, ligne_debut_cible=ligne_debut_cible_auto),
        ColonneMapping("Suivi des arrêts", 33, col_source=54, ligne_debut_source=11, ligne_debut_cible=ligne_debut_cible_auto),
        ColonneMapping("Suivi des arrêts", 46, col_source=79, ligne_debut_source=11, ligne_debut_cible=ligne_debut_cible_auto),
        ColonneMapping("Suivi des arrêts", 9, col_source=2, ligne_debut_source=11, ligne_debut_cible=ligne_debut_cible_auto),
        ColonneMapping("Suivi des arrêts", 22, col_source=28, ligne_debut_source=11, ligne_debut_cible=ligne_debut_cible_auto),
        ColonneMapping("Suivi des arrêts", 35, col_source=53, ligne_debut_source=11, ligne_debut_cible=ligne_debut_cible_auto),
        ColonneMapping("Suivi des arrêts", 48, col_source=78, ligne_debut_source=11, ligne_debut_cible=ligne_debut_cible_auto),
        # Sommes Temps d'arrêt N process
        ColonneMapping("Suivi des arrêts", 8, colonnes_source=[4, 5, 6, 7, 8, 9], type="somme", ligne_debut_source=11, ligne_debut_cible=ligne_debut_cible_auto),
        ColonneMapping("Suivi des arrêts", 21, colonnes_source=[30, 31, 32, 33, 34, 35], type="somme", ligne_debut_source=11, ligne_debut_cible=ligne_debut_cible_auto),
        ColonneMapping("Suivi des arrêts", 34, colonnes_source=[55, 56, 57, 58, 59, 60], type="somme", ligne_debut_source=11, ligne_debut_cible=ligne_debut_cible_auto),
        ColonneMapping("Suivi des arrêts", 47, colonnes_source=[80, 81, 82, 83, 84, 85], type="somme", ligne_debut_source=11, ligne_debut_cible=ligne_debut_cible_auto),
        # Sommes Temps d'arrêts Non planifié Maintenance
        ColonneMapping("Suivi des arrêts", 10, colonnes_source=[10, 11, 12, 13], type="somme", ligne_debut_source=11, ligne_debut_cible=ligne_debut_cible_auto),
        ColonneMapping("Suivi des arrêts", 23, colonnes_source=[36, 37, 38, 39], type="somme", ligne_debut_source=11, ligne_debut_cible=ligne_debut_cible_auto),
        ColonneMapping("Suivi des arrêts", 36, colonnes_source=[61, 62, 63, 64], type="somme", ligne_debut_source=11, ligne_debut_cible=ligne_debut_cible_auto),
        ColonneMapping("Suivi des arrêts", 49, colonnes_source=[86, 87, 88, 89], type="somme", ligne_debut_source=11, ligne_debut_cible=ligne_debut_cible_auto),
        # Sommes Temps d'arrêts externes (h)
        ColonneMapping("Suivi des arrêts", 11, colonnes_source=[14, 15, 16, 17, 18, 19, 20, 21, 22, 23], type="somme", ligne_debut_source=11, ligne_debut_cible=ligne_debut_cible_auto),
        ColonneMapping("Suivi des arrêts", 24, colonnes_source=[40, 41, 42, 43, 44, 45, 46, 47, 48, 49], type="somme", ligne_debut_source=11, ligne_debut_cible=ligne_debut_cible_auto),
        ColonneMapping("Suivi des arrêts", 37, colonnes_source=[65, 66, 67, 68, 69, 70, 71, 72, 73, 74], type="somme", ligne_debut_source=11, ligne_debut_cible=ligne_debut_cible_auto),
        ColonneMapping("Suivi des arrêts", 50, colonnes_source=[90, 91, 92, 93, 94, 95, 96, 97, 98, 99], type="somme", ligne_debut_source=11, ligne_debut_cible=ligne_debut_cible_auto),
    ]

    # Trouver les positions des sauts dans le premier élément de Synthèse
    premier_element = colonnes_a_copier[0]
    ws_source = wb_source[premier_element.feuille_source]
    col_source = premier_element.col_source
    ligne_debut_source = premier_element.ligne_debut_source

    # Détecter automatiquement la dernière date valide dans le fichier source
    derniere_date_source, derniere_ligne_date_source, dernier_total_ligne_source = detecter_derniere_date_source(ws_source, col_source, ligne_debut_source)

    if derniere_date_source:
        print(f"Dernière date détectée automatiquement dans le fichier source : {derniere_date_source}")
    else:
        print("Aucune date trouvée dans le fichier source !")
        return None

    # --- Génération du plan de recopie (dates + index de ligne source) ---
    print(f"Recherche de la première date valide...")
    premiere_date = None
    premiere_ligne = None
    for row in range(ligne_debut_source, dernier_total_ligne_source):
        val = ws_source.cell(row=row, column=col_source).value
        if isinstance(val, datetime):
            premiere_date = val
            premiere_ligne = row
            print(f"Première date trouvée: {premiere_date.strftime('%d/%m/%Y')} (ligne {row})")
            break
    if premiere_date is None:
        print("Aucune date trouvée !")
        return None

    # Générer la liste des lignes à traiter (en sautant Total et la ligne suivante)
    lignes_a_traiter = []
    ligne_courante = premiere_ligne
    date_courante = premiere_date
    while ligne_courante < dernier_total_ligne_source:
        val = ws_source.cell(row=ligne_courante, column=col_source).value
        if isinstance(val, str) and val == "Total":
            print(f"  Saut de la ligne {ligne_courante} (Total) et {ligne_courante+1} (après Total)")
            ligne_courante += 2
            continue
        if val is None or str(val).strip() == "":
            print(f"  Saut de la ligne {ligne_courante} (vide)")
            ligne_courante += 1
            continue
        lignes_a_traiter.append((ligne_courante, date_courante.strftime("%d/%m/%Y")))
        date_courante = date_courante + timedelta(days=1)
        ligne_courante += 1

    print(f"\nPlan de recopie généré ({len(lignes_a_traiter)} lignes) :")
    for idx, (ligne_src, date_str) in enumerate(lignes_a_traiter):
        print(f"  {date_str} (index {idx}, ligne source {ligne_src})")
    print(f"De {lignes_a_traiter[0][1]} à {lignes_a_traiter[-1][1]}")

    # --- Recopie des données dans le fichier cible ---
    row_cible = ligne_debut_cible_auto
    for i, (ligne_src, date_str) in enumerate(lignes_a_traiter):
        print(f"\nRecopie pour la date {date_str} (ligne source {ligne_src}, index {i}) dans la ligne cible {row_cible}")
        valeurs_ligne = {}
        for mapping in colonnes_a_copier:
            col_cible = mapping.col_cible
            if col_cible == 1:
                continue
            feuille = wb_source[mapping.feuille_source]
            if mapping.feuille_source == "Synthèse":
                ligne_source = ligne_src
            else:  # Suivi des arrêts
                # Décaler l'index pour la feuille Suivi des arrêts (commence à 11)
                offset = ligne_src - 10  # 10 est le début de Synthèse
                ligne_source = 11 + offset
            if mapping.type == "somme":
                somme = 0
                for col_source in mapping.colonnes_source:
                    v = feuille.cell(row=ligne_source, column=col_source).value
                    print(f"    Somme: feuille={mapping.feuille_source}, ligne={ligne_source}, col={col_source}, val={v}")
                    if v is not None and str(v).strip() != "":
                        try:
                            somme += float(v)
                        except ValueError:
                            pass
                valeurs_ligne[col_cible] = somme
                print(f"    -> Somme totale pour col {col_cible}: {somme}")
            else:
                col_source = mapping.col_source
                v = feuille.cell(row=ligne_source, column=col_source).value
                print(f"    Copie: feuille={mapping.feuille_source}, ligne={ligne_source}, col={col_source}, val={v}")
                valeurs_ligne[col_cible] = v
        ws_cible.cell(row=row_cible, column=1).value = date_str
        for col_cible, valeur in valeurs_ligne.items():
            ws_cible.cell(row=row_cible, column=col_cible).value = valeur
        row_cible += 1

    print(f"\nNombre de lignes ajoutées : {len(lignes_a_traiter)}")
    print(f"Dernière ligne cible utilisée : {row_cible - 1}")


    fichier_sortie = 'Fichier Prétraité MP1.xlsx'
    wb_cible.save(fichier_sortie)
    print("✅ ---------------------------------------------success-------------------------------------------------------------")
    return fichier_sortie

if __name__ == "__main__":
    # Exemple d'utilisation de la fonction
    fichier_source = 'Fichier source MP1.xlsx'
    fichier_cible = 'Fichier cible MP1.xlsx'
    
    fichier_resultat = traiter_fichiers_MP1(fichier_source, fichier_cible)
    print(f"Fichier de sortie généré : {fichier_resultat}")
