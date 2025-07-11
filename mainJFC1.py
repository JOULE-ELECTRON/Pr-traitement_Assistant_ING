import openpyxl
from openpyxl.styles import Alignment
from datetime import datetime
import os

class Mapping:
    def __init__(self, source_col, target_col, name):
        self.source_col = source_col
        self.target_col = target_col
        self.name = name

class PretraitementProcessor:
    def __init__(self, chemin_source, chemin_cible):
        self.wb_source = openpyxl.load_workbook(chemin_source, data_only=True)
        self.ws_source = self.wb_source.active
        self.source_max_row = self.ws_source.max_row
        self.wb_cible = openpyxl.load_workbook(chemin_cible)
        self.ws_cible = self.wb_cible.active
        # Mapping des colonnes (1-based)
        self.date_col_source = 4
        self.date_col_cible = 1
        self.production_mapping = Mapping(5, 6, "Production")
        self.arret_mappings = [
            Mapping(35, 8, "Temps d'arrêts planifiés process (h)"),
            Mapping(36, 7, "Temps d'arrêts planifiés maintenance (h)"),
            Mapping(37, 9, "Temps grande révision (h)"),
            Mapping(38, 11, "Temps d'arrêts non planifiés process (h)"),
            Mapping(39, 10, "Temps d'arrêts non planifiés maintenance (h)"),
            Mapping(40, 12, "Temps d'arrêts externes (h)")
        ]
        self.titre_mappings = [
            Mapping(10, 13, "%P2O5 ACP28% produit"),
            Mapping(14, 14, "%P2O5 ACP54% produit")
        ]

    def percent_str(self, val):
        try:
            return f"{round(float(val), 2)}%" if val is not None else None
        except Exception:
            return f"{val}%" if val is not None else None

    def get_source_cell(self, row, col):
        return self.ws_source.cell(row=row, column=col).value

    def trouver_derniere_ligne_date(self):
        derniere_ligne = 1
        for row in range(2, self.ws_cible.max_row + 1):
            valeur = self.ws_cible.cell(row=row, column=self.date_col_cible).value
            if valeur is not None and str(valeur).strip() != "":
                try:
                    if isinstance(valeur, datetime):
                        derniere_ligne = row
                    elif isinstance(valeur, str) and len(valeur.split('/')) == 3:
                        datetime.strptime(valeur, "%d/%m/%Y")
                        derniere_ligne = row
                except ValueError:
                    pass
        return derniere_ligne + 1

    def collecter_dates_existantes(self):
        dates_existantes = set()
        for row in range(2, self.ws_cible.max_row + 1):
            valeur = self.ws_cible.cell(row=row, column=self.date_col_cible).value
            if valeur is not None and str(valeur).strip() != "":
                try:
                    if isinstance(valeur, datetime):
                        dates_existantes.add(valeur.strftime("%d/%m/%Y"))
                    elif isinstance(valeur, str) and len(valeur.split('/')) == 3:
                        datetime.strptime(valeur, "%d/%m/%Y")
                        dates_existantes.add(valeur)
                except ValueError:
                    pass
        return dates_existantes

    def traiter(self):
        ligne_debut = self.trouver_derniere_ligne_date()
        dates_existantes = self.collecter_dates_existantes()
        dates_a_ajouter = []  # (row_source, date_str)
        max_lignes = 200  # Limite à 200 lignes pour accélérer les tests (à retirer si besoin)
        derniere_ligne_source = min(self.source_max_row, max_lignes + 1)  # +1 car on commence à 2
        for row in range(2, derniere_ligne_source):
            date_val = self.get_source_cell(row, self.date_col_source)
            if date_val is None or str(date_val).strip() == "":
                continue
            try:
                if isinstance(date_val, datetime):
                    date_str = date_val.strftime("%d/%m/%Y")
                else:
                    date_val_str = str(date_val).strip()
                    if len(date_val_str.split('/')) == 3:
                        date_val = datetime.strptime(date_val_str, "%d/%m/%Y")
                        date_str = date_val.strftime("%d/%m/%Y")
                    else:
                        continue
                if date_str not in dates_existantes:
                    dates_a_ajouter.append((row, date_str))
            except Exception:
                continue
        current_row = ligne_debut
        for row_source, date_str in dates_a_ajouter:
            # 1. Date
            cell = self.ws_cible.cell(row=current_row, column=self.date_col_cible)
            cell.value = date_str
            cell.alignment = Alignment(horizontal='center', vertical='center')
            # 2. Production
            prod_value = self.get_source_cell(row_source, self.production_mapping.source_col)
            cell = self.ws_cible.cell(row=current_row, column=self.production_mapping.target_col)
            cell.value = prod_value
            cell.alignment = Alignment(horizontal='center', vertical='center')
            # 3. Arrêts
            for arret_map in self.arret_mappings:
                arret_value = self.get_source_cell(row_source, arret_map.source_col)
                cell = self.ws_cible.cell(row=current_row, column=arret_map.target_col)
                cell.value = arret_value
                cell.alignment = Alignment(horizontal='center', vertical='center')
            # 4. Titres (pourcentages et centrage)
            titre_ee = self.get_source_cell(row_source, 10)
            titre_se = self.get_source_cell(row_source, 14)
            percent_ee = self.percent_str(titre_ee)
            percent_se = self.percent_str(titre_se)
            values = [(13, percent_ee), (14, percent_se)]
            for col, val in values:
                cell = self.ws_cible.cell(row=current_row, column=col)
                cell.value = val
                cell.alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1
        fichier_sortie = 'Fichier Prétraité Jorf.xlsx'
        self.wb_cible.save(fichier_sortie)
        print("\n✅ ----------------------------------------------------SUCCES--------------------------------------------")
        return fichier_sortie

def traiter_fichiers_pretraitement(chemin_source, chemin_cible):
    processor = PretraitementProcessor(chemin_source, chemin_cible)
    return processor.traiter()

if __name__ == "__main__":
    fichier_source = 'Fichier Source JFC1_Mod.xlsx'  # Utiliser un fichier Excel par défaut
    fichier_cible = 'Fichier Cible Jorf.xlsx'
    fichier_resultat = traiter_fichiers_pretraitement(fichier_source, fichier_cible)
    print(f"Fichier de sortie généré : {fichier_resultat}")